import argparse
import importlib.util
import json
import math
import re
import zipfile
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


AUTO_SHEET_DEFAULT = "Suggestion Export"
HA_SHEET_DEFAULT = "Stock"
PO_TOOL_DIR = Path(__file__).resolve().parents[1]
PO_TOOL_SCRIPT = PO_TOOL_DIR / "po-offer" / "generate_po_offer.py"
XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def parse_number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    text = re.sub(r"\([A-Z]{3}\)$", "", text).strip()
    if text.endswith("%"):
        text = text[:-1].strip()
        try:
            return float(text) / 100
        except ValueError:
            return 0.0
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_decimal_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u00a0", "").replace(" ", "")
    text = re.sub(r"[^0-9,.\-()]", "", text)
    if not text:
        return None
    is_negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        pieces = text.split(",")
        if len(pieces[-1]) in (1, 2):
            text = "".join(pieces[:-1]) + "." + pieces[-1]
        else:
            text = text.replace(",", "")
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    return -number if is_negative else number


def parse_optional_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return parse_number(value)


def parse_margin(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number / 100 if abs(number) > 1 else number
    text = str(value).strip()
    if text.endswith("%"):
        return parse_number(text)
    number = parse_number(text)
    return number / 100 if abs(number) > 1 else number


def is_existing_product(value):
    return normalize_header(value) == "existing product"


def as_identifier(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    try:
        num = float(text)
        if num.is_integer():
            return str(int(num))
    except ValueError:
        pass
    return text


def first_present(*values):
    for value in values:
        if value is not None and value != "":
            return value
    return None


def calculated_margin(cost, promo_price):
    if cost in (None, 0) or promo_price is None:
        return None
    return (promo_price - cost) / cost


def decimal_to_float(value):
    if value is None:
        return None
    return float(value)


def normalize_lookup_key(value):
    text = as_identifier(value)
    return re.sub(r"\s+", "", text).upper()


def find_col(header_map, *names, required=True):
    for name in names:
        key = normalize_header(name)
        if key in header_map:
            return header_map[key]
    if required:
        raise ValueError(f"Missing required column: {' / '.join(names)}")
    return None


def build_header_map(ws, row_number):
    return {
        normalize_header(cell.value): cell.column
        for cell in ws[row_number]
        if normalize_header(cell.value)
    }


def get_cell(row, col):
    if not col:
        return None
    if col - 1 >= len(row):
        return None
    return row[col - 1]


def excel_col_index(name):
    value = 0
    for char in name:
        if "A" <= char <= "Z":
            value = value * 26 + ord(char) - 64
    return value


def xml_cell_column(ref):
    match = re.match(r"([A-Z]+)", ref or "")
    return match.group(1) if match else ""


def xml_read_shared_strings(zf):
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    strings = []
    for si in root.findall("main:si", XML_NS):
        pieces = [node.text or "" for node in si.findall(".//main:t", XML_NS)]
        strings.append("".join(pieces))
    return strings


def xml_read_cell_value(cell, shared_strings):
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//main:t", XML_NS))
    value = cell.find("main:v", XML_NS)
    if value is None:
        return None
    raw = value.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    return raw


def xml_find_sheet_path(zf, sheet_name):
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("pkgrel:Relationship", XML_NS)
    }
    first_path = None
    for sheet in workbook.findall("main:sheets/main:sheet", XML_NS):
        rel_id = sheet.attrib[f"{{{XML_NS['rel']}}}id"]
        target = rel_map[rel_id].lstrip("/")
        path = target if target.startswith("xl/") else f"xl/{target}"
        if first_path is None:
            first_path = path
        if sheet.attrib.get("name") == sheet_name:
            return path
    if first_path:
        return first_path
    raise ValueError("Workbook has no sheets.")


def find_workbook_sheet(wb, preferred_name):
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    return wb[wb.sheetnames[0]]


def extract_auto_rows(auto_path, supplier_code=None):
    wb = load_workbook(auto_path, read_only=True, data_only=True)
    ws = find_workbook_sheet(wb, AUTO_SHEET_DEFAULT)
    headers = build_header_map(ws, 1)

    sku_col = find_col(headers, "SKU #")
    name_col = find_col(headers, "SKU name")
    ean_col = find_col(headers, "UPC/EAN", "UPC/EAN Barcode", "Barcode")
    barcode_col = find_col(headers, "Barcode", required=False)
    vendor_col = find_col(headers, "Vendor#")
    qty_col = find_col(headers, "Replenish Qty")
    new_product_col = find_col(headers, "Is New Product")
    margin_col = find_col(headers, "MDSE Margin (7 days)", "Gross margin (7 days)", required=False)
    inv_cost_col = find_col(headers, "Inventory unit cost (Including VAT)", required=False)
    list_price_col = find_col(headers, "List price", required=False)
    promo_price_col = find_col(headers, "Promotion price", required=False)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if supplier_code and str(get_cell(row, vendor_col) or "").strip() != supplier_code:
            continue
        if not is_existing_product(get_cell(row, new_product_col)):
            continue
        qty = parse_number(get_cell(row, qty_col))
        if qty == 0:
            continue
        ean = as_identifier(get_cell(row, ean_col)) or as_identifier(get_cell(row, barcode_col))
        rows.append({
            "sku": as_identifier(get_cell(row, sku_col)),
            "sku_name": str(get_cell(row, name_col) or "").strip(),
            "upc_ean_barcode": ean,
            "inquiry_qty": qty,
            "margin": parse_margin(get_cell(row, margin_col)),
            "long_aging": None,
            "non_active_90d": None,
            "source": "Auto PO suggestion",
            "auto_vendor": str(get_cell(row, vendor_col) or "").strip(),
            "inv_price_incl_vat": None,
            "auto_inv_price_incl_vat": parse_optional_number(get_cell(row, inv_cost_col)),
            "auto_jd_price": parse_optional_number(get_cell(row, list_price_col)),
            "auto_promo_price": parse_optional_number(get_cell(row, promo_price_col)),
            "need_rp": None,
            "rp_blacklist": None,
            "exclude_reason": "",
        })
    return rows


def build_ha_records(ha_path):
    with zipfile.ZipFile(ha_path) as zf:
        shared_strings = xml_read_shared_strings(zf)
        sheet_path = xml_find_sheet_path(zf, HA_SHEET_DEFAULT)
        header_map = {}
        max_col = 0
        header_row = None

        for _, elem in ET.iterparse(zf.open(sheet_path), events=("end",)):
            if elem.tag != f"{{{XML_NS['main']}}}row":
                continue
            if int(elem.attrib.get("r", "0")) == 2:
                header_row = elem
                break
            elem.clear()

        if header_row is None:
            raise ValueError("Could not find HA sheet header row 2.")
        for cell in header_row.findall("main:c", XML_NS):
            col_name = xml_cell_column(cell.attrib.get("r", ""))
            col_idx = excel_col_index(col_name)
            value = xml_read_cell_value(cell, shared_strings)
            if normalize_header(value):
                header_map[normalize_header(value)] = col_idx
            max_col = max(max_col, col_idx)

    cols = {
        "long_aging": find_col(header_map, "Long-aging?", required=False),
        "non_active_90d": find_col(header_map, "Non-active in 90D?", required=False),
        "sku": find_col(header_map, "SKU"),
        "ean": find_col(header_map, "EAN", required=False),
        "sku_name": find_col(header_map, "SKU Name"),
        "supplier_no": find_col(header_map, "Supplier #"),
        "last_supplier": find_col(header_map, "Last Supplier", required=False),
        "last_supplier_no": find_col(header_map, "Last Supplier #", required=False),
        "category_manager": find_col(header_map, "Category Manager", required=False),
        "inv_price": find_col(header_map, "Inv Price (incl. VAT)"),
        "jd_price": find_col(header_map, "JD Price", required=False),
        "promo_price": find_col(header_map, "Promo Price", required=False),
        "margin": find_col(header_map, "Margin"),
        "need_rp": find_col(header_map, "Need RP?"),
        "suggested_qty": find_col(header_map, "Suggested RP QTY"),
        "rp_blacklist": find_col(header_map, "RP Blacklist?", required=False),
        "reasons": find_col(header_map, "Reasons", required=False),
        "remarks": find_col(header_map, "Remarks", required=False),
    }
    az_col = 52 if max_col >= 52 else cols["reasons"]

    records = {}
    all_rows = []
    needed_indices = set(cols.values()) | {az_col}
    needed_indices.discard(None)

    with zipfile.ZipFile(ha_path) as zf:
        shared_strings = xml_read_shared_strings(zf)
        sheet_path = xml_find_sheet_path(zf, HA_SHEET_DEFAULT)
        for _, elem in ET.iterparse(zf.open(sheet_path), events=("end",)):
            if elem.tag != f"{{{XML_NS['main']}}}row":
                continue
            if int(elem.attrib.get("r", "0")) < 3:
                elem.clear()
                continue
            values = {}
            for cell in elem.findall("main:c", XML_NS):
                col_name = xml_cell_column(cell.attrib.get("r", ""))
                col_idx = excel_col_index(col_name)
                if col_idx in needed_indices:
                    values[col_idx] = xml_read_cell_value(cell, shared_strings)

            sku = as_identifier(values.get(cols["sku"]))
            if not sku:
                elem.clear()
                continue
            reason_parts = [
                str(values.get(az_col) or ""),
                str(values.get(cols["reasons"]) or ""),
                str(values.get(cols["remarks"]) or ""),
            ]
            record = {
                "sku": sku,
                "sku_name": str(values.get(cols["sku_name"]) or "").strip(),
                "upc_ean_barcode": as_identifier(values.get(cols["ean"])),
                "supplier_no": str(values.get(cols["supplier_no"]) or "").strip(),
                "last_supplier": str(values.get(cols["last_supplier"]) or "").strip(),
                "last_supplier_no": str(values.get(cols["last_supplier_no"]) or "").strip(),
                "category_manager": str(values.get(cols["category_manager"]) or "").strip(),
                "inv_price_incl_vat": parse_number(values.get(cols["inv_price"])),
                "jd_price": parse_number(values.get(cols["jd_price"])),
                "promo_price": parse_number(values.get(cols["promo_price"])),
                "margin": parse_margin(values.get(cols["margin"])),
                "need_rp": str(values.get(cols["need_rp"]) or "").strip(),
                "suggested_rp_qty": parse_number(values.get(cols["suggested_qty"])),
                "rp_blacklist": str(values.get(cols["rp_blacklist"]) or "").strip().upper(),
                "long_aging": str(values.get(cols["long_aging"]) or "").strip(),
                "non_active_90d": str(values.get(cols["non_active_90d"]) or "").strip(),
                "reason_text": " | ".join(part for part in reason_parts if part.strip()),
            }
            existing = records.get(sku)
            if existing is None or "vendor eol" in record["reason_text"].lower():
                records[sku] = record
            all_rows.append(record)
            elem.clear()
    return records, all_rows


def should_exclude_vendor_eol(record):
    return "vendor eol" in (record.get("reason_text") or "").lower()


def supplier_code_matches(value, supplier_code):
    text = str(value or "").strip()
    target = str(supplier_code or "").strip()
    if not text or not target:
        return False
    vendor_codes = re.findall(r"Vendor_\d+", text, flags=re.IGNORECASE)
    if vendor_codes:
        return any(code.upper() == target.upper() for code in vendor_codes)
    if text.upper() == target.upper():
        return True
    target_suffix = re.sub(r"^VENDOR[_ -]?", "", target.upper())
    return bool(target_suffix and target_suffix != target.upper() and text.upper() == target_suffix)


def supplier_code_contained(value, supplier_code):
    """Match a supplier code appearing anywhere in a Supplier # cell."""
    text = str(value or "").strip().upper()
    target = str(supplier_code or "").strip().upper()
    if not text or not target:
        return False
    if target in text:
        return True
    # Some HA exports store Supplier # as only the numeric suffix (000443),
    # while the workbench input uses the canonical Vendor_000443 form.
    target_suffix = re.sub(r"^VENDOR[_ -]?", "", target)
    return bool(target_suffix and target_suffix != target and target_suffix in text)


def is_blank_supplier_value(value):
    return str(value or "").strip().upper() in {"", "NULL", "NONE", "NAN", "N/A", "-"}


def apply_price_priority(row, ha_by_sku, auto_by_sku):
    sku = row.get("sku")
    ha = ha_by_sku.get(sku, {})
    auto = auto_by_sku.get(sku, {})
    cost = first_present(auto.get("auto_inv_price_incl_vat"), ha.get("inv_price_incl_vat"), row.get("inv_price_incl_vat"))
    jd_price = first_present(auto.get("auto_jd_price"), ha.get("jd_price"), row.get("jd_price"))
    promo_price = first_present(auto.get("auto_promo_price"), ha.get("promo_price"), row.get("promo_price"))
    row["inv_price_incl_vat"] = cost
    row["jd_price"] = jd_price
    row["promo_price"] = promo_price
    row["margin"] = calculated_margin(cost, promo_price)
    return row


def create_supplier_inventory_template_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier Inventory"
    headers = ["EAN", "Available Stock", "Price", "Case Pack Qty"]
    notes = [
        "必填：请填写询单表里的 UPC/EAN Barcode",
        "可选：不填表示库存充足",
        "可选：不填则该 SKU 不生成 offer",
        "可选：不填或填 0 表示没有箱规；填写大于 0 的数值后 PO 数量向下取整到箱规整倍数",
    ]
    ws.append(headers)
    ws.append(notes)
    ws.append(["示例：8720389013720", "", "", ""])
    widths = {"A": 24, "B": 18, "C": 14, "D": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor="EAF2F8")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A3"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def read_supplier_inventory(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = build_header_map(ws, 1)
    ean_col = find_col(headers, "EAN", "UPC/EAN", "UPC/EAN Barcode", "Barcode")
    sku_col = find_col(headers, "SKU", "SKU #", required=False)
    stock_col = find_col(headers, "Available Stock", "Available inventory", "Stock", "库存", "可用库存", required=False)
    price_col = find_col(headers, "Price", "price-excl", "报价", "价格", required=False)
    case_pack_col = find_col(headers, "Case Pack Qty", "Case Pack", "CTN qty", "箱规", required=False)

    records = {}
    for row_idx in range(2, ws.max_row + 1):
        ean = as_identifier(ws.cell(row_idx, ean_col).value)
        sku = as_identifier(ws.cell(row_idx, sku_col).value) if sku_col else ""
        if not ean and not sku:
            continue
        record = {
            "row": row_idx,
            "ean": ean,
            "sku": sku,
            "stock": parse_decimal_value(ws.cell(row_idx, stock_col).value) if stock_col else None,
            "price": parse_decimal_value(ws.cell(row_idx, price_col).value) if price_col else None,
            "case_pack": parse_decimal_value(ws.cell(row_idx, case_pack_col).value) if case_pack_col else None,
        }
        if ean:
            records[("ean", normalize_lookup_key(ean))] = record
        if sku:
            records[("sku", normalize_lookup_key(sku))] = record
    return records


def match_supplier_inventory(record_map, item):
    ean = normalize_lookup_key(item.get("upc_ean_barcode"))
    sku = normalize_lookup_key(item.get("sku"))
    return record_map.get(("ean", ean)) or record_map.get(("sku", sku))


def supplier_offer_rate(supplier_code):
    path = PO_TOOL_DIR / "supplier offer information.xlsx.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb["OPC-Bulk offer"] if "OPC-Bulk offer" in wb.sheetnames else wb.worksheets[0]
    headers = build_header_map(ws, 2)
    vendor_col = find_col(headers, "* Vendor #", "Vendor #", "Vendor#")
    rate_col = find_col(headers, "* Rate-input", "Rate-input", required=False)
    if not rate_col:
        return Decimal("0")
    for row_idx in range(3, ws.max_row + 1):
        vendor = as_identifier(ws.cell(row_idx, vendor_col).value)
        if vendor.upper() == supplier_code.upper():
            rate = parse_decimal_value(ws.cell(row_idx, rate_col).value)
            return rate or Decimal("0")
    raise ValueError(f"supplier offer information 中找不到供应商编码：{supplier_code}")


def final_po_quantity(inquiry_qty, stock, case_pack):
    inquiry = parse_decimal_value(inquiry_qty)
    if inquiry is None or inquiry <= 0:
        return 0
    available = inquiry if stock is None else stock
    if available <= 0:
        return 0
    qty = min(inquiry, available)
    qty_int = math.floor(float(qty))
    if qty_int <= 0:
        return 0
    if case_pack is not None and case_pack > 0:
        pack = math.floor(float(case_pack))
        if pack > 0:
            qty_int = (qty_int // pack) * pack
    return qty_int


def load_po_tool_module():
    if not PO_TOOL_SCRIPT.exists():
        raise FileNotFoundError(f"PO/Offer reference tool not found: {PO_TOOL_SCRIPT}")
    spec = importlib.util.spec_from_file_location("po_offer_generator", PO_TOOL_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def build_po_input_from_inquiry(output_dir, supplier_code, rows, inventory_path, timestamp):
    inventory = read_supplier_inventory(inventory_path)
    po_input_path = output_dir / f"{supplier_code}_PO_Offer_Input_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Offer Input"
    ws.append(["SKU", "EAN", "Vendor Code", "Quantity", "price-excl", "PO type"])

    included = 0
    skipped = []
    for item in rows:
        if item.get("po_offer_eligible") is False:
            skipped.append({
                "sku": item.get("sku"),
                "ean": item.get("upc_ean_barcode"),
                "reason": item.get("po_offer_skip_reason") or "Not eligible for PO/offer",
            })
            continue
        inv = match_supplier_inventory(inventory, item)
        if not inv:
            skipped.append({
                "sku": item.get("sku"),
                "ean": item.get("upc_ean_barcode"),
                "reason": "Supplier inventory row not found for this SKU/EAN",
            })
            continue
        stock = inv.get("stock") if inv else None
        price = inv.get("price") if inv else None
        case_pack = inv.get("case_pack") if inv else None
        quantity = final_po_quantity(item.get("inquiry_qty"), stock, case_pack)
        if quantity <= 0:
            skipped.append({
                "sku": item.get("sku"),
                "ean": item.get("upc_ean_barcode"),
                "reason": "PO quantity is 0 after stock/case pack calculation",
            })
            continue
        ws.append([
            item.get("sku"),
            item.get("upc_ean_barcode"),
            supplier_code,
            quantity,
            float(price) if price is not None else None,
            "Regular",
        ])
        included += 1

    for col, width in {"A": 15, "B": 22, "C": 18, "D": 12, "E": 12, "F": 12}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(po_input_path)
    return po_input_path, included, skipped


def generate_po_offer_files(output_dir, supplier_code, rows, inventory_path, timestamp):
    po_input_path, included, skipped = build_po_input_from_inquiry(
        output_dir, supplier_code, rows, inventory_path, timestamp
    )
    po_module = load_po_tool_module()
    old_output_dir = po_module.OUTPUT_DIR
    po_module.OUTPUT_DIR = output_dir
    stamp = f"{supplier_code}_PO_Offer_{timestamp}"
    run_output_dir = output_dir / stamp
    try:
        offer_path, po_path, error_path = po_module.build_outputs(
            input_path=po_input_path,
            supplier_offer_path=PO_TOOL_DIR / "supplier offer information.xlsx.xlsx",
            supplier_po_path=PO_TOOL_DIR / "supplier PO information.xlsx.xlsx",
            site_product_path=PO_TOOL_DIR / "Export via custom template_460893_20260708044154.xlsx",
            run_stamp=stamp,
        )
    except PermissionError as exc:
        offer_path = run_output_dir / "generated_OPC_bulk_offer.xlsx"
        if not offer_path.exists():
            offer_path = None
        po_path = run_output_dir / "generated_Bulk_Import_PO.xlsx"
        if not po_path.exists():
            raise
        error_path = run_output_dir / f"error_report_fallback_{datetime.now().strftime('%H%M%S')}.xlsx"
        error_wb = Workbook()
        ws = error_wb.active
        ws.title = "Errors"
        ws.append(["input_row", "sku", "ean", "vendor", "errors"])
        ws.append(["", "", "", supplier_code, f"Default error_report.xlsx could not be saved: {exc}"])
        error_wb.save(error_path)
    finally:
        po_module.OUTPUT_DIR = old_output_dir

    if skipped:
        if not error_path.exists():
            error_wb = Workbook()
            ws = error_wb.active
            ws.title = "Errors"
            ws.append(["input_row", "sku", "ean", "vendor", "errors"])
        else:
            error_wb = load_workbook(error_path)
        ws = error_wb.active
        for item in skipped:
            ws.append(["", item["sku"], item["ean"], supplier_code, item["reason"]])
        try:
            error_wb.save(error_path)
        except PermissionError:
            error_path = error_path.with_name(f"{error_path.stem}_fallback_{datetime.now().strftime('%H%M%S')}.xlsx")
            error_wb.save(error_path)

    errors = po_module.read_error_rows(error_path) if hasattr(po_module, "read_error_rows") else []

    return {
        "po_input_path": str(po_input_path.resolve()),
        "po_path": str(po_path.resolve()),
        "offer_path": str(offer_path.resolve()) if offer_path else None,
        "error_path": str(error_path.resolve()),
        "errors": errors,
        "po_item_count": included,
    }


def make_ha_item(ha):
    return {
        "sku": ha["sku"],
        "sku_name": ha["sku_name"],
        "upc_ean_barcode": ha["upc_ean_barcode"],
        "inquiry_qty": ha["suggested_rp_qty"],
        "margin": ha["margin"],
        "long_aging": ha["long_aging"],
        "non_active_90d": ha["non_active_90d"],
        "source": "HA master sheet",
        "inv_price_incl_vat": ha["inv_price_incl_vat"],
        "jd_price": ha.get("jd_price"),
        "promo_price": ha.get("promo_price"),
        "last_supplier": ha.get("last_supplier"),
        "last_supplier_no": ha.get("last_supplier_no"),
        "category_manager": ha.get("category_manager"),
        "need_rp": ha["need_rp"],
        "rp_blacklist": ha["rp_blacklist"],
        "exclude_reason": "",
    }


def eligible_ha_replenishment(ha):
    if ha["inv_price_incl_vat"] == 0:
        return False
    if str(ha.get("need_rp") or "").strip().upper() != "Y":
        return False
    if ha["suggested_rp_qty"] <= 0:
        return False
    if should_exclude_vendor_eol(ha):
        return False
    if ha.get("rp_blacklist") == "Y":
        return False
    return True


def inventory_key_set(inventory):
    return {
        key_value
        for key_type, key_value in inventory
        if key_type in {"ean", "sku"} and key_value
    }


def inventory_matches_item(inventory, item):
    return match_supplier_inventory(inventory, item) is not None


def build_full_inventory_rows(auto_path, ha_by_sku, ha_rows, supplier_code, inventory_path):
    inventory = read_supplier_inventory(inventory_path)
    tax_multiplier = Decimal("1.21")

    all_auto_rows = extract_auto_rows(auto_path, None)
    auto_by_sku = {row["sku"]: row for row in all_auto_rows}
    candidates = {}
    stats = {
        "auto_count": 0,
        "ha_candidate_count": 0,
        "excluded_blacklist": 0,
        "excluded_vendor_eol": 0,
        "price_rejected": 0,
        "last_supplier_match": 0,
        "last_supplier_price_match": 0,
        "price_not_qualified_inquiry_only": 0,
    }

    for item in all_auto_rows:
        if not inventory_matches_item(inventory, item):
            continue
        ha = ha_by_sku.get(item["sku"])
        if ha:
            if ha.get("rp_blacklist") == "Y":
                stats["excluded_blacklist"] += 1
                continue
            if should_exclude_vendor_eol(ha):
                stats["excluded_vendor_eol"] += 1
                continue
            item["last_supplier"] = ha.get("last_supplier")
            item["last_supplier_no"] = ha.get("last_supplier_no")
            item["category_manager"] = ha.get("category_manager")
            item["long_aging"] = ha.get("long_aging")
            item["non_active_90d"] = ha.get("non_active_90d")
            item["rp_blacklist"] = ha.get("rp_blacklist")
            if not item.get("upc_ean_barcode"):
                item["upc_ean_barcode"] = ha.get("upc_ean_barcode")
            if not item.get("sku_name"):
                item["sku_name"] = ha.get("sku_name")
        apply_price_priority(item, ha_by_sku, auto_by_sku)
        candidates.setdefault(item["sku"], item)
        stats["auto_count"] += 1

    for ha in ha_rows:
        if not eligible_ha_replenishment(ha):
            continue
        item = make_ha_item(ha)
        if not inventory_matches_item(inventory, item):
            continue
        apply_price_priority(item, ha_by_sku, auto_by_sku)
        candidates.setdefault(item["sku"], item)
        stats["ha_candidate_count"] += 1

    combined = []
    for item in candidates.values():
        inv = match_supplier_inventory(inventory, item)
        stock_price_excl = inv.get("price") if inv else None
        stock_price_incl = stock_price_excl * tax_multiplier if stock_price_excl is not None else None
        current_cost = parse_decimal_value(item.get("inv_price_incl_vat"))
        same_last_supplier = supplier_code_matches(item.get("last_supplier_no"), supplier_code)

        item["supplier_inventory_price_excl"] = decimal_to_float(stock_price_excl)
        item["supplier_inventory_tax_incl_price"] = decimal_to_float(stock_price_incl)
        item["original_last_supplier"] = item.get("last_supplier")
        item["original_last_supplier_no"] = item.get("last_supplier_no")
        item["full_match_note"] = "Last supplier matched" if same_last_supplier else "Price check pending"
        item["po_offer_eligible"] = False
        item["po_offer_skip_reason"] = ""

        if stock_price_incl is not None and current_cost not in (None, Decimal("0")):
            diff = stock_price_incl - current_cost
            item["supplier_price_diff"] = decimal_to_float(diff)
            item["supplier_price_diff_rate"] = decimal_to_float(diff / current_cost)
        else:
            item["supplier_price_diff"] = None
            item["supplier_price_diff_rate"] = None

        if same_last_supplier:
            stats["last_supplier_match"] += 1
            item["po_offer_eligible"] = True
            combined.append(item)
            continue

        if stock_price_incl is not None and current_cost is not None and stock_price_incl <= current_cost:
            stats["last_supplier_price_match"] += 1
            item["po_offer_eligible"] = True
            item["full_match_note"] = "Price qualified for PO/offer"
            combined.append(item)
        elif stock_price_incl is not None and current_cost is not None and stock_price_incl > current_cost:
            stats["price_not_qualified_inquiry_only"] += 1
            item["full_match_note"] = "Inquiry only: supplier tax-incl price > inventory cost"
            item["po_offer_skip_reason"] = "Supplier tax-incl price > Inventory unit cost"
            combined.append(item)
        else:
            stats["price_rejected"] += 1

    stats["final_count"] = len(combined)
    stats["total_inquiry_qty"] = sum(parse_number(row["inquiry_qty"]) for row in combined)
    return combined, auto_by_sku, stats


def generate(auto_path, ha_path, supplier_code, output_dir, inventory_path=None, mode="vendor"):
    auto_path = Path(auto_path)
    ha_path = Path(ha_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # The web UI uses `inventory` for the full inventory matching option,
    # while the generator's internal name is `full_inventory`.
    if mode in {"inventory", "full-inventory", "full_inventory"}:
        mode = "full_inventory"
    else:
        mode = "vendor"

    ha_by_sku, ha_rows = build_ha_records(ha_path)

    if mode == "full_inventory":
        if not inventory_path:
            raise ValueError("全量匹配供应商库存表模式需要上传供应商库存表。")
        combined, auto_by_sku, mode_stats = build_full_inventory_rows(
            auto_path, ha_by_sku, ha_rows, supplier_code, inventory_path
        )
        auto_rows = []
        ha_candidates = []
        excluded_blacklist = mode_stats["excluded_blacklist"]
        excluded_vendor_eol = mode_stats["excluded_vendor_eol"]
    else:
        auto_rows = extract_auto_rows(auto_path, supplier_code)
        auto_by_sku = {row["sku"]: row for row in auto_rows}

        combined = []
        seen = set()
        excluded_blacklist = 0
        excluded_vendor_eol = 0

        for item in auto_rows:
            sku = item["sku"]
            ha = ha_by_sku.get(sku)
            if ha:
                if ha.get("rp_blacklist") == "Y":
                    excluded_blacklist += 1
                    continue
                if should_exclude_vendor_eol(ha):
                    excluded_vendor_eol += 1
                    continue
                item["inv_price_incl_vat"] = ha.get("inv_price_incl_vat")
                item["jd_price"] = ha.get("jd_price")
                item["promo_price"] = ha.get("promo_price")
                item["last_supplier"] = ha.get("last_supplier")
                item["last_supplier_no"] = ha.get("last_supplier_no")
                item["category_manager"] = ha.get("category_manager")
                item["long_aging"] = ha.get("long_aging")
                item["non_active_90d"] = ha.get("non_active_90d")
                item["rp_blacklist"] = ha.get("rp_blacklist")
                if ha.get("margin") is not None:
                    item["margin"] = ha.get("margin")
                if not item.get("upc_ean_barcode"):
                    item["upc_ean_barcode"] = ha.get("upc_ean_barcode")
                if not item.get("sku_name"):
                    item["sku_name"] = ha.get("sku_name")
            combined.append(item)
            seen.add(sku)

        ha_candidates = []
        for ha in ha_rows:
            last_supplier_no = str(ha.get("last_supplier_no") or "").strip()
            matches_last_supplier = supplier_code_matches(last_supplier_no, supplier_code)
            matches_supplier_column_without_last = (
                is_blank_supplier_value(last_supplier_no)
                and supplier_code_contained(ha.get("supplier_no"), supplier_code)
            )
            if not (matches_last_supplier or matches_supplier_column_without_last):
                continue
            if not eligible_ha_replenishment(ha):
                if should_exclude_vendor_eol(ha):
                    excluded_vendor_eol += 1
                elif ha.get("rp_blacklist") == "Y":
                    excluded_blacklist += 1
                continue
            ha_candidates.append(ha)
            if ha["sku"] in seen:
                continue
            combined.append(make_ha_item(ha))
            seen.add(ha["sku"])

        for row in combined:
            apply_price_priority(row, ha_by_sku, auto_by_sku)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{supplier_code}_Inquiry_SKU_List_Final_{timestamp}.xlsx"
    stats_payload = {
        "auto_count": mode_stats["auto_count"] if mode == "full_inventory" else len(auto_rows),
        "ha_candidate_count": mode_stats["ha_candidate_count"] if mode == "full_inventory" else len(ha_candidates),
        "final_count": len(combined),
        "excluded_blacklist": excluded_blacklist,
        "excluded_vendor_eol": excluded_vendor_eol,
        "total_inquiry_qty": sum(parse_number(row["inquiry_qty"]) for row in combined),
        "price_rejected": mode_stats.get("price_rejected", 0) if mode == "full_inventory" else 0,
        "last_supplier_match": mode_stats.get("last_supplier_match", 0) if mode == "full_inventory" else 0,
        "last_supplier_price_match": mode_stats.get("last_supplier_price_match", 0) if mode == "full_inventory" else 0,
        "price_not_qualified_inquiry_only": mode_stats.get("price_not_qualified_inquiry_only", 0) if mode == "full_inventory" else 0,
        "mode": mode,
    }
    write_workbook(output_path, supplier_code, auto_path, ha_path, combined, ha_by_sku, auto_by_sku, stats_payload)
    result = {
        "output_path": str(output_path.resolve()),
        "supplier_code": supplier_code,
        "auto_count": stats_payload["auto_count"],
        "ha_candidate_count": stats_payload["ha_candidate_count"],
        "final_count": len(combined),
        "excluded_blacklist": excluded_blacklist,
        "excluded_vendor_eol": excluded_vendor_eol,
        "price_rejected": stats_payload["price_rejected"],
        "last_supplier_match": stats_payload["last_supplier_match"],
        "last_supplier_price_match": stats_payload["last_supplier_price_match"],
        "price_not_qualified_inquiry_only": stats_payload["price_not_qualified_inquiry_only"],
        "mode": mode,
        "total_inquiry_qty": sum(parse_number(row["inquiry_qty"]) for row in combined),
    }
    if inventory_path:
        result["po_offer"] = generate_po_offer_files(output_dir, supplier_code, combined, inventory_path, timestamp)
    return result


def build_low_margin_rows(rows, ha_by_sku, auto_by_sku):
    low_margin_rows = []
    seen = set()
    for item in rows:
        sku = item.get("sku")
        if not sku or sku in seen:
            continue
        margin = item.get("margin")
        if margin is None or margin >= 0.05:
            continue
        seen.add(sku)
        ha = ha_by_sku.get(sku)
        auto = auto_by_sku.get(sku, {})
        if ha:
            low_margin_rows.append({
                "sku": sku,
                "ean": ha.get("upc_ean_barcode") or item.get("upc_ean_barcode"),
                "sku_name": ha.get("sku_name") or item.get("sku_name"),
                "inv_price_incl_vat": item.get("inv_price_incl_vat"),
                "jd_price": item.get("jd_price"),
                "promo_price": item.get("promo_price"),
                "margin": item.get("margin"),
                "last_supplier": ha.get("last_supplier"),
                "last_supplier_no": ha.get("last_supplier_no"),
                "category_manager": ha.get("category_manager"),
                "long_aging": ha.get("long_aging"),
            })
        else:
            low_margin_rows.append({
                "sku": sku,
                "ean": auto.get("upc_ean_barcode") or item.get("upc_ean_barcode"),
                "sku_name": auto.get("sku_name") or item.get("sku_name"),
                "inv_price_incl_vat": item.get("inv_price_incl_vat"),
                "jd_price": item.get("jd_price"),
                "promo_price": item.get("promo_price"),
                "margin": item.get("margin"),
                "last_supplier": "",
                "last_supplier_no": "",
                "category_manager": "",
                "long_aging": item.get("long_aging"),
            })
    return low_margin_rows


def write_workbook(output_path, supplier_code, auto_path, ha_path, rows, ha_by_sku, auto_by_sku, stats):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiry List"
    ws.sheet_view.showGridLines = False
    is_full_inventory = stats.get("mode") == "full_inventory"

    headers = [
        "SKU #",
        "SKU name",
        "UPC/EAN Barcode",
        "Inquiry Qty",
        "Inventory unit cost (Including VAT)",
        "Promotion price",
        "Margin",
        "Long-aging?",
        "Non-active in 90D?",
        "Source",
    ]
    if is_full_inventory:
        headers.extend([
            "Original Last Supplier",
            "Original Last Supplier #",
            "Supplier Inventory Tax-incl Price",
            "Price Difference",
            "Difference Rate",
            "PO/Offer Eligibility",
        ])

    last_data_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_data_col}1")
    ws["A1"] = f"{supplier_code} Inquiry SKU List"
    ws.merge_cells(f"A2:{last_data_col}2")
    ws["A2"] = f"Table A: {auto_path} | Table B: {ha_path}"
    ws.append([])
    ws.append(headers)

    for item in rows:
        row_values = [
            item.get("sku"),
            item.get("sku_name"),
            item.get("upc_ean_barcode"),
            item.get("inquiry_qty"),
            item.get("inv_price_incl_vat"),
            item.get("promo_price"),
            item.get("margin"),
            item.get("long_aging"),
            item.get("non_active_90d"),
            item.get("source"),
        ]
        if is_full_inventory:
            row_values.extend([
                item.get("original_last_supplier"),
                item.get("original_last_supplier_no"),
                item.get("supplier_inventory_tax_incl_price"),
                item.get("supplier_price_diff"),
                item.get("supplier_price_diff_rate"),
                "Yes" if item.get("po_offer_eligible") is not False else item.get("po_offer_skip_reason"),
            ])
        ws.append(row_values)

    metric_start = 1
    metric_col = len(headers) + 2
    metric_value_col = len(headers) + 3
    ws.cell(metric_start, metric_col, "Metric")
    ws.cell(metric_start, metric_value_col, "Value")
    metrics = [
        ("Auto PO SKUs", stats["auto_count"]),
        ("HA Candidate SKUs", stats["ha_candidate_count"]),
        ("Excluded RP Blacklist", stats["excluded_blacklist"]),
        ("Excluded Vendor EOL", stats.get("excluded_vendor_eol", 0)),
        ("Price Rejected", stats.get("price_rejected", 0)),
        ("Last Supplier Matched", stats.get("last_supplier_match", 0)),
        ("Price Qualified", stats.get("last_supplier_price_match", 0)),
        ("Inquiry Only Price Not Qualified", stats.get("price_not_qualified_inquiry_only", 0)),
        ("Final SKU Count", stats["final_count"]),
        ("Total Inquiry Qty", stats["total_inquiry_qty"]),
    ]
    for idx, (name, value) in enumerate(metrics, start=2):
        ws.cell(idx, metric_col, name)
        ws.cell(idx, metric_value_col, value)

    blue = "1F4E79"
    light_blue = "EAF2F8"
    header_blue = "5B9BD5"
    green = "70AD47"
    red_fill = "F4CCCC"
    thin = Side(style="thin", color="D9E2F3")

    ws["A1"].fill = PatternFill("solid", fgColor=blue)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = PatternFill("solid", fgColor=light_blue)
    ws["A2"].font = Font(color=blue, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        if cell.column <= len(headers):
            cell.fill = PatternFill("solid", fgColor=header_blue)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

    for cell in (ws.cell(1, metric_col), ws.cell(1, metric_value_col)):
        cell.fill = PatternFill("solid", fgColor=green)
        cell.font = Font(bold=True, color="FFFFFF")

    widths = {
        "A": 15, "B": 64, "C": 20, "D": 13, "E": 24,
        "F": 15, "G": 11, "H": 13, "I": 18, "J": 20,
        "K": 28, "L": 18, "M": 24, "N": 16, "O": 14,
        get_column_letter(metric_col): 22, get_column_letter(metric_value_col): 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=4, max_row=max(ws.max_row, 4), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column in (2,))

    for row_idx in range(5, ws.max_row + 1):
        ws.cell(row_idx, 1).number_format = "@"
        ws.cell(row_idx, 3).number_format = "0"
        ws.cell(row_idx, 4).number_format = "#,##0.0"
        ws.cell(row_idx, 5).number_format = "#,##0.00"
        ws.cell(row_idx, 6).number_format = "#,##0.00"
        ws.cell(row_idx, 7).number_format = "0.0%"
        margin = ws.cell(row_idx, 7).value
        if margin is not None and margin < 0.05:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=red_fill)
        if is_full_inventory:
            ws.cell(row_idx, 13).number_format = "#,##0.00"
            ws.cell(row_idx, 14).number_format = "#,##0.00"
            ws.cell(row_idx, 15).number_format = "0.0%"

    for row_idx in range(2, 10):
        ws.cell(row_idx, metric_value_col).number_format = "#,##0.0" if row_idx == 9 else "#,##0"

    if rows:
        table_ref = f"A4:{last_data_col}{ws.max_row}"
        table = Table(displayName="InquiryList", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    ws.freeze_panes = "A5"

    low_rows = build_low_margin_rows(rows, ha_by_sku, auto_by_sku)
    low_ws = wb.create_sheet("low margin& OOS sku")
    low_ws.sheet_view.showGridLines = False
    low_headers = [
        "SKU",
        "EAN",
        "SKU Name",
        "Inventory unit cost (Including VAT)",
        "JD Price",
        "Promotion price",
        "Margin",
        "Last Supplier",
        "Last Supplier #",
        "Category Manager",
        "Long-aging?",
    ]
    low_ws.append(low_headers)
    for item in low_rows:
        low_ws.append([
            item.get("sku"),
            item.get("ean"),
            item.get("sku_name"),
            item.get("inv_price_incl_vat"),
            item.get("jd_price"),
            item.get("promo_price"),
            item.get("margin"),
            item.get("last_supplier"),
            item.get("last_supplier_no"),
            item.get("category_manager"),
            item.get("long_aging"),
        ])

    for cell in low_ws[1]:
        cell.fill = PatternFill("solid", fgColor=header_blue)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    low_widths = {
        "A": 15, "B": 20, "C": 64, "D": 24, "E": 15,
        "F": 15, "G": 11, "H": 28, "I": 18, "J": 20, "K": 13,
    }
    for col, width in low_widths.items():
        low_ws.column_dimensions[col].width = width

    for row in low_ws.iter_rows(min_row=2, max_row=max(low_ws.max_row, 2), min_col=1, max_col=len(low_headers)):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column in (3, 8, 10))

    for row_idx in range(2, low_ws.max_row + 1):
        low_ws.cell(row_idx, 1).number_format = "@"
        low_ws.cell(row_idx, 2).number_format = "0"
        low_ws.cell(row_idx, 4).number_format = "#,##0.00"
        low_ws.cell(row_idx, 5).number_format = "#,##0.00"
        low_ws.cell(row_idx, 6).number_format = "#,##0.00"
        low_ws.cell(row_idx, 7).number_format = "0.0%"
        for col_idx in range(1, len(low_headers) + 1):
            low_ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=red_fill)

    if low_rows:
        low_table = Table(displayName="LowMarginOOSList", ref=f"A1:K{low_ws.max_row}")
        low_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        low_ws.add_table(low_table)
    low_ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate vendor replenishment inquiry workbook.")
    parser.add_argument("--auto", required=True, help="Path to Auto PO suggestion workbook")
    parser.add_argument("--ha", required=True, help="Path to HA master sheet workbook")
    parser.add_argument("--supplier", required=True, help="Supplier code, for example Vendor_002756")
    parser.add_argument("--output-dir", default="outputs/inquiry_tool", help="Directory for generated workbook")
    parser.add_argument("--inventory", help="Optional supplier inventory workbook for PO/offer generation")
    parser.add_argument("--mode", choices=["vendor", "full_inventory"], default="vendor", help="Generation mode")
    args = parser.parse_args()
    result = generate(args.auto, args.ha, args.supplier, args.output_dir, args.inventory, args.mode)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
