from __future__ import annotations

import argparse
import glob
import json
import os
import re
import sys
from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_BASE_DIR = Path(__file__).resolve().parents[1]
BASE_DIR = Path(os.getenv("PO_TOOL_HOME", DEFAULT_BASE_DIR)).resolve()

OFFER_TEMPLATE = BASE_DIR / "OPC-bulk-offer.xlsx"
PO_TEMPLATE = BASE_DIR / "Bulk Import PO.xlsx"
SUPPLIER_OFFER = BASE_DIR / "supplier offer information.xlsx.xlsx"
SUPPLIER_PO = BASE_DIR / "supplier PO information.xlsx.xlsx"
OPC_SITE_PRODUCT = BASE_DIR / "Export via custom template_460893_20260708044154.xlsx"
OUTPUT_DIR = Path(os.getenv("PO_OUTPUT_DIR", BASE_DIR / "outputs")).resolve()
CONFIG_PATH = Path(os.getenv("PO_CONFIG_PATH", BASE_DIR / ".po_offer_config.json")).resolve()

WAREHOUSE_CODE_MAP = {
    "_15_Small item": "NLRM01",
    "_14_Medium item": "NLRM02",
    "_2_Large item": "NLRM04",
}

INPUT_HEADER_ALIASES = {
    "sku": {"sku", "sku#", "sku #", "* sku", "商品编码", "商品编号"},
    "ean": {"ean", "upc", "barcode", "条码", "ean码"},
    "vendor": {
        "vendor",
        "vendor #",
        "vendor#",
        "vendor code",
        "vendor_code",
        "* vendor code",
        "* vendor code ",
        "供应商",
        "供应商编码",
        "供应商编号",
    },
    "quantity": {"quantity", "qty", "* quantity", "数量", "采购数量"},
    "price_excl": {
        "price-excl",
        "price excl",
        "price_excl",
        "price excl.",
        "不含税价",
        "税前价",
        "价格",
    },
    "po_type": {"po type", "* po type", "po_type", "po类型", "po type类型"},
    "original_po": {"original po", "original_po", "原po", "原 po", "原始po"},
}


def normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\n", " ")
    return re.sub(r"\s+", " ", text).lower()


def normalize_key(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def find_latest_site_product_file() -> Path:
    patterns = [
        "OPC_site product*.xlsx",
        "OPC_site product*.xlsm",
        "Export via custom template*.xlsx",
        "Export via custom template*.xlsm",
    ]
    matches = []
    for pattern in patterns:
        matches.extend(glob.glob(str(BASE_DIR / pattern)))
    matches = sorted(set(matches), reverse=True)
    if not matches:
        raise FileNotFoundError("未找到 OPC_site product 文件")
    return Path(matches[0])


def resolve_path(value: str | Path | None) -> Path | None:
    if value in (None, ""):
        return None
    path = Path(str(value).strip().strip('"'))
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve()


def config_value(path: Path) -> str:
    try:
        return str(path.relative_to(BASE_DIR))
    except ValueError:
        return str(path)


def load_config() -> dict[str, str]:
    if not CONFIG_PATH.exists():
        return {}
    try:
        with CONFIG_PATH.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return {}


def save_config(config: dict[str, str]) -> None:
    with CONFIG_PATH.open("w", encoding="utf-8") as fh:
        json.dump(config, fh, ensure_ascii=False, indent=2)


def existing_path(path: Path, label: str) -> Path:
    if not path.exists():
        raise FileNotFoundError(f"{label} file does not exist: {path}")
    return path


def optional_config_path(config: dict[str, str], key: str) -> Path | None:
    path = resolve_path(config.get(key))
    if path and path.exists():
        return path
    return None


def choose_reference_path(
    provided: str | None,
    config: dict[str, str],
    config_key: str,
    default_path: Path | None,
    label: str,
) -> Path | None:
    provided_path = resolve_path(provided)
    if provided_path:
        return existing_path(provided_path, label)

    configured_path = optional_config_path(config, config_key)
    if configured_path:
        return configured_path

    if default_path:
        return existing_path(default_path, label)

    return None


def get_reference_paths(
    supplier_offer: str | None = None,
    supplier_po: str | None = None,
    site_product: str | None = None,
    input_path: str | None = None,
) -> tuple[Path | None, Path, Path, Path]:
    config = load_config()

    default_site_product = None
    try:
        default_site_product = find_latest_site_product_file()
    except FileNotFoundError:
        pass

    resolved_input = choose_reference_path(input_path, config, "last_input", None, "Input")
    resolved_offer = choose_reference_path(
        supplier_offer, config, "supplier_offer", SUPPLIER_OFFER, "Supplier offer information"
    )
    resolved_po = choose_reference_path(supplier_po, config, "supplier_po", SUPPLIER_PO, "Supplier PO information")
    resolved_site_product = choose_reference_path(
        site_product,
        config,
        "site_product",
        default_site_product or OPC_SITE_PRODUCT,
        "OPC_site product",
    )

    if not resolved_offer:
        raise FileNotFoundError("Supplier offer information file is required.")
    if not resolved_po:
        raise FileNotFoundError("Supplier PO information file is required.")
    if not resolved_site_product:
        raise FileNotFoundError("OPC_site product file is required.")

    config.update(
        {
            "supplier_offer": config_value(resolved_offer),
            "supplier_po": config_value(resolved_po),
            "site_product": config_value(resolved_site_product),
        }
    )
    config.pop("replenish", None)
    if resolved_input:
        config["last_input"] = config_value(resolved_input)
    save_config(config)

    return resolved_input, resolved_offer, resolved_po, resolved_site_product


def worksheet_headers(ws, header_row=2) -> dict[str, int]:
    headers = {}
    for cell in ws[header_row]:
        name = normalize_header(cell.value)
        if name:
            headers[name] = cell.column
    return headers


def col(headers: dict[str, int], name: str) -> int:
    key = normalize_header(name)
    if key not in headers:
        raise KeyError(f"找不到列：{name}")
    return headers[key]


def read_supplier_map(path: Path, sheet_name: str | None, vendor_header: str) -> dict[str, dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    headers = worksheet_headers(ws)
    vendor_col = col(headers, vendor_header)
    header_by_col = {idx: name for name, idx in headers.items()}
    result = {}
    for row in range(3, ws.max_row + 1):
        vendor = normalize_key(ws.cell(row, vendor_col).value)
        if not vendor:
            continue
        result[vendor] = {
            header_by_col[c]: ws.cell(row, c).value
            for c in range(1, ws.max_column + 1)
            if c in header_by_col
        }
    return result


def split_codes(value) -> list[str]:
    text = normalize_key(value)
    if not text:
        return []
    return [normalize_key(part) for part in re.split(r"[,，;；\s]+", text) if normalize_key(part)]


def read_site_product(path: Path) -> tuple[dict[str, str], dict[str, str], dict[str, str], dict[str, list[str]]]:
    sku_to_warehouse = {}
    upc_ean_to_sku = {}
    barcode_to_sku = {}
    sku_to_eans = {}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header_row = None
    headers = {}
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        row_headers = {
            normalize_header(cell.value): cell.column
            for cell in ws[row_idx]
            if normalize_header(cell.value)
        }
        if "sku id" in row_headers and "upc/ean code" in row_headers and "barcode" in row_headers:
            header_row = row_idx
            headers = row_headers
            break
    if not header_row:
        raise KeyError("OPC_site product 表缺少表头：SKU ID / UPC/EAN Code / Barcode")

    sku_col = headers["sku id"]
    upc_col = headers["upc/ean code"]
    barcode_col = headers["barcode"]
    warehouse_col = headers.get("warehouse_name") or headers.get("item size")
    if not warehouse_col:
        raise KeyError("OPC_site product 表缺少 warehouse_name 或 Item size 列")

    for row in ws.iter_rows(min_row=header_row + 1):
        sku = normalize_key(row[sku_col - 1].value)
        if not sku:
            continue
        warehouse = normalize_key(row[warehouse_col - 1].value)
        if warehouse and sku not in sku_to_warehouse:
            sku_to_warehouse[sku] = warehouse
        for code in split_codes(row[upc_col - 1].value):
            upc_ean_to_sku[code] = sku
            sku_to_eans.setdefault(sku, []).append(code)
        for code in split_codes(row[barcode_col - 1].value):
            barcode_to_sku[code] = sku

    return upc_ean_to_sku, barcode_to_sku, sku_to_warehouse, sku_to_eans


def read_input_rows(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    raw_headers = [normalize_header(cell.value) for cell in ws[1]]
    alias_to_field = {}
    for field, aliases in INPUT_HEADER_ALIASES.items():
        for alias in aliases:
            alias_to_field[normalize_header(alias)] = field

    field_cols = {}
    for idx, header in enumerate(raw_headers, start=1):
        if header in alias_to_field:
            field_cols[alias_to_field[header]] = idx

    required_any = ("sku", "ean")
    if not any(name in field_cols for name in required_any):
        raise KeyError("输入表需要至少包含 SKU 或 EAN 列")
    for name in ("vendor", "quantity"):
        if name not in field_cols:
            raise KeyError(f"输入表缺少必填列：{name}")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        item = {"_row": row_idx}
        has_value = False
        for field, col_idx in field_cols.items():
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                has_value = True
            item[field] = value
        if has_value:
            rows.append(item)
    return rows


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        return value
    text = str(value).strip().replace(",", "")
    try:
        number = Decimal(text)
    except InvalidOperation:
        return value
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def parse_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u00a0", "").replace(" ", "")
    text = re.sub(r"[^0-9,.\-()]", "", text)
    if not text:
        return None

    is_negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts[-1]) in (1, 2):
            text = "".join(parts[:-1]) + "." + parts[-1]
        else:
            text = text.replace(",", "")

    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    return -number if is_negative else number


def parse_price_excl(value):
    number = parse_decimal(value)
    if number is None:
        return None
    return float(number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def parse_quantity(value):
    number = parse_decimal(value)
    if number is None:
        return None
    return int(number.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_data_rows(ws, start_row=3) -> None:
    for row in range(start_row, ws.max_row + 1):
        for cell in ws[row]:
            cell.value = None


def write_by_header(ws, headers: dict[str, int], row: int, header: str, value) -> None:
    ws.cell(row, col(headers, header)).value = value


def format_by_header(ws, headers: dict[str, int], row: int, header: str, number_format: str) -> None:
    ws.cell(row, col(headers, header)).number_format = number_format


def build_outputs(
    input_path: Path,
    supplier_offer_path: Path,
    supplier_po_path: Path,
    site_product_path: Path,
    run_stamp: str | None = None,
) -> tuple[Path | None, Path, Path]:
    run_stamp = run_stamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = OUTPUT_DIR / run_stamp
    output_dir.mkdir(parents=True, exist_ok=True)

    upc_ean_to_sku, barcode_to_sku, sku_to_warehouse, _ = read_site_product(site_product_path)
    offer_suppliers = read_supplier_map(supplier_offer_path, "OPC-Bulk offer", "* Vendor #")
    po_suppliers = read_supplier_map(supplier_po_path, "Sheet1", "* Vendor Code ")
    input_rows = read_input_rows(input_path)

    offer_wb = load_workbook(OFFER_TEMPLATE)
    offer_ws = offer_wb["OPC-Bulk offer"] if "OPC-Bulk offer" in offer_wb.sheetnames else offer_wb.worksheets[0]
    offer_headers = worksheet_headers(offer_ws)
    clear_data_rows(offer_ws)

    po_wb = load_workbook(PO_TEMPLATE)
    po_ws = po_wb.worksheets[0]
    po_headers = worksheet_headers(po_ws)
    clear_data_rows(po_ws)

    errors = []
    offer_row = 3
    po_row = 3
    offer_count = 0
    for source in input_rows:
        raw_sku = normalize_key(source.get("sku"))
        ean = normalize_key(source.get("ean"))
        vendor = normalize_key(source.get("vendor"))
        sku = raw_sku or upc_ean_to_sku.get(ean, "") or barcode_to_sku.get(ean, "")
        quantity = parse_quantity(source.get("quantity"))
        price_excl = parse_price_excl(source.get("price_excl"))
        has_price = price_excl not in (None, "")
        po_type = normalize_key(source.get("po_type")) or "Regular"
        original_po = normalize_key(source.get("original_po"))

        row_errors = []
        if not sku:
            row_errors.append(f"无法通过 UPC/EAN Code 或 Barcode 找到 SKU：{ean}")
        if has_price and vendor not in offer_suppliers:
            row_errors.append(f"offer 供应商信息未找到：{vendor}")
        if vendor not in po_suppliers:
            row_errors.append(f"PO 供应商信息未找到：{vendor}")
        warehouse_name = sku_to_warehouse.get(sku, "")
        warehouse_code = WAREHOUSE_CODE_MAP.get(warehouse_name, "")
        if po_type == "Abnormal-PO" and not original_po:
            row_errors.append("PO type 为 Abnormal-PO 时必须填写 Original PO")
        if quantity in (None, ""):
            row_errors.append("Quantity 为空")

        if row_errors:
            errors.append(
                {
                    "input_row": source["_row"],
                    "sku": sku,
                    "ean": ean,
                    "vendor": vendor,
                    "errors": "; ".join(row_errors),
                }
            )
            continue

        if has_price:
            copy_row_style(offer_ws, 3, offer_row, offer_ws.max_column)
            offer_info = offer_suppliers[vendor]
            for header in (
                "* Purchase entity",
                "* Vendor #",
                "* Incoterm",
                "* Currency",
                "* Rate-input",
                "Site",
                "case pack Qty",
                "Case units\n1-piece\n2-kilogram\n3-bottle",
            ):
                key = normalize_header(header)
                if key in offer_info and key in offer_headers:
                    write_by_header(offer_ws, offer_headers, offer_row, header, offer_info[key])
            write_by_header(offer_ws, offer_headers, offer_row, "SKU #", sku)
            write_by_header(offer_ws, offer_headers, offer_row, "price-excl", price_excl)
            format_by_header(offer_ws, offer_headers, offer_row, "price-excl", "0.00")
            offer_row += 1
            offer_count += 1

        copy_row_style(po_ws, 3, po_row, po_ws.max_column)
        po_info = po_suppliers[vendor]
        for header in (
            "* Vendor Code ",
            "* Purchase entity",
            "* Site",
            "* Currency",
            "* Incoterm",
        ):
            key = normalize_header(header)
            if key in po_info and key in po_headers:
                write_by_header(po_ws, po_headers, po_row, header, po_info[key])
        write_by_header(po_ws, po_headers, po_row, "* SKU", sku)
        write_by_header(po_ws, po_headers, po_row, "* Warehouse Code", warehouse_code)
        write_by_header(po_ws, po_headers, po_row, "* Quantity", quantity)
        format_by_header(po_ws, po_headers, po_row, "* Quantity", "0")
        write_by_header(po_ws, po_headers, po_row, "* PO type", po_type)
        if original_po:
            write_by_header(po_ws, po_headers, po_row, "Original PO", original_po)

        po_row += 1

    offer_path = output_dir / "generated_OPC_bulk_offer.xlsx" if offer_count else None
    po_path = output_dir / "generated_Bulk_Import_PO.xlsx"
    error_path = output_dir / "error_report.xlsx"
    if offer_path:
        offer_wb.save(offer_path)
    po_wb.save(po_path)

    error_wb = Workbook()
    ws = error_wb.active
    ws.title = "Errors"
    ws.append(["input_row", "sku", "ean", "vendor", "errors"])
    for err in errors:
        ws.append([err["input_row"], err["sku"], err["ean"], err["vendor"], err["errors"]])
    ws.freeze_panes = "A2"
    error_wb.save(error_path)

    return offer_path, po_path, error_path


def read_error_rows(error_path):
    """Read generated error-report rows for API clients that render them inline."""
    if not error_path or not Path(error_path).exists():
        return []
    wb = load_workbook(error_path, read_only=True, data_only=True)
    ws = wb.active
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(value or "").strip() for value in header_values]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({headers[idx]: values[idx] for idx in range(min(len(headers), len(values))) if headers[idx]})
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate OPC offer and Bulk Import PO files.")
    parser.add_argument("input", nargs="?", help="输入表路径；不填时使用上次输入过的文件")
    parser.add_argument("--stamp", help="输出文件夹名称后缀；默认使用当前时间")
    parser.add_argument("--supplier-offer", help="新的 supplier offer information 文件路径；不填时使用上次保存的文件")
    parser.add_argument("--supplier-po", help="新的 supplier PO information 文件路径；不填时使用上次保存的文件")
    parser.add_argument("--site-product", help="新的 OPC_site product 文件路径；不填时使用上次保存的文件")
    parser.add_argument("--replenish", help=argparse.SUPPRESS)
    parser.add_argument("--show-config", action="store_true", help="显示当前保存的默认文件路径")
    args = parser.parse_args()

    try:
        if args.show_config:
            config = load_config()
            print(json.dumps(config, ensure_ascii=False, indent=2))
            return 0

        input_path, supplier_offer_path, supplier_po_path, site_product_path = get_reference_paths(
            supplier_offer=args.supplier_offer or os.getenv("PO_SUPPLIER_OFFER"),
            supplier_po=args.supplier_po or os.getenv("PO_SUPPLIER_PO"),
            site_product=args.site_product or args.replenish or os.getenv("PO_SITE_PRODUCT") or os.getenv("PO_REPLENISH"),
            input_path=args.input,
        )
        if not input_path:
            raise ValueError("Input file is required for the first run. Later runs can leave it blank.")

        offer_path, po_path, error_path = build_outputs(
            input_path=input_path,
            supplier_offer_path=supplier_offer_path,
            supplier_po_path=supplier_po_path,
            site_product_path=site_product_path,
            run_stamp=args.stamp,
        )
    except Exception as exc:
        print(f"生成失败：{exc}", file=sys.stderr)
        return 1

    print(f"Input: {input_path}")
    print(f"Supplier offer information: {supplier_offer_path}")
    print(f"Supplier PO information: {supplier_po_path}")
    print(f"OPC_site product: {site_product_path}")
    print(f"Offer: {offer_path if offer_path else 'Skipped because no price-excl was provided'}")
    print(f"PO: {po_path}")
    print(f"Error report: {error_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
